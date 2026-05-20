"""
Ingere a planilha do Diogo (Fechamento ONLOG.xlsx) e gera onlog_diff.json,
que e' inlinado no dashboard como ONLOG_DIFF e mostra a conferencia da quinzena.

Workflow toda quinzena:
    1. Receber a planilha do Diogo
    2. py ingest_diogo_onlog.py "C:\\caminho\\Fechamento ONLOG.xlsx"
    3. py merge_data.py && py build_html.py
    4. git add -A && git commit -m "Onlog: conferencia <quinzena>" && git push

Saida:
    onlog_diff.json
        {
            "quinzena": {"de": "...", "ate": "..."},
            "geradoEm": "...",
            "planilhaArquivo": "Fechamento ONLOG.xlsx",
            "resumo": {"ok": N, "divergencias": N, "soPlanilha": N, "soFabric": N,
                       "nPlanilha": N, "nFabric": N},
            "divergencias": [{"orderNumber":..., "marca":..., "campo":..., "planilha":..., "fabric":...}, ...],
            "soPlanilha":   [{"codigoVolume":..., "orderNumber":..., "cliente":..., "cidade":..., "uf":..., "status":..., "postagem":...}, ...],
            "soFabric":     [{"data":..., "orderNumber":..., "marca":..., "cliente":..., "cidade":..., "uf":..., "cancelado":..., "status":..., "valor":...}, ...]
        }
"""

import argparse
import json
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERRO: openpyxl nao instalado. Rode: py -m pip install openpyxl")
    sys.exit(1)

ROOT = Path(__file__).parent
ONLOG_JSON = ROOT / "onlog_data.json"


def norm_txt(s) -> str:
    if s is None:
        return ""
    nfd = unicodedata.normalize("NFD", str(s))
    no_diac = "".join(c for c in nfd if unicodedata.category(c) != "Mn")
    return " ".join("".join(c if c.isalnum() else " " for c in no_diac.upper()).split())


def norm_uf(s) -> str:
    return (str(s or "").strip().upper())[:2]


def parse_val_br(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def read_planilha(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        return []
    header = [str(h).strip() if h else "" for h in rows_raw[0]]
    out = []
    for r in rows_raw[1:]:
        out.append({header[i]: r[i] for i in range(len(header))})
    return out


def detect_quinzenas(rows: list[dict]) -> list[tuple[str, str]]:
    """Retorna TODAS as quinzenas (de, ate) cobertas pelas datas da planilha.
    Permite ingerir planilhas que cobrem mais de uma quinzena (ex: mes inteiro)."""
    halves = set()
    for r in rows:
        d = r.get("Data")
        s = ""
        if isinstance(d, datetime):
            s = d.date().isoformat()
        elif isinstance(d, str):
            s = d[:10]
        if not s:
            continue
        try:
            y, mo, dia = int(s[:4]), int(s[5:7]), int(s[8:10])
        except ValueError:
            continue
        halves.add((y, mo, 1 if dia <= 15 else 2))
    from calendar import monthrange
    out = []
    for y, mo, half in sorted(halves):
        mes = f"{y:04d}-{mo:02d}"
        if half == 1:
            out.append((f"{mes}-01", f"{mes}-15"))
        else:
            out.append((f"{mes}-16", f"{mes}-{monthrange(y, mo)[1]:02d}"))
    return out


def detect_quinzena(rows: list[dict]) -> tuple[str, str]:
    min_d = None
    for r in rows:
        d = r.get("Data")
        s = ""
        if isinstance(d, datetime):
            s = d.date().isoformat()
        elif isinstance(d, str):
            s = d[:10]
        if not s:
            continue
        if min_d is None or s < min_d:
            min_d = s
    if not min_d:
        return "", ""
    y, mo, dia = int(min_d[:4]), int(min_d[5:7]), int(min_d[8:10])
    mes = min_d[:7]
    if dia <= 15:
        return f"{mes}-01", f"{mes}-15"
    from calendar import monthrange
    return f"{mes}-16", f"{mes}-{monthrange(y, mo)[1]:02d}"


def detect_diogo_total(rows: list[dict]) -> float | None:
    """Detecta a linha totalizadora da planilha do Diogo:
    todos os campos vazios EXCETO ValorPostagem. Esse e' o valor que ele cobra."""
    for r in rows:
        if r.get("Destinatario") or r.get("CodigoInterno") or r.get("CodigoVolume"):
            continue
        if r.get("Remetente") or r.get("NumeroPedido") or r.get("NumeroNF"):
            continue
        v = parse_val_br(r.get("ValorPostagem"))
        if v is not None and v > 100:  # totalizador sempre tem valor alto
            return v
    return None


def aggregate_planilha(rows: list[dict], de: str, ate: str) -> tuple[dict, list]:
    """Retorna (pedidos_por_codigovolume, pa_vesti_avulsas).

    PA VESTI = linhas sem CodigoVolume (sem NumeroPedido) - postagens manuais
    geradas pela equipe Vesti direto no painel Onlog/Jadlog.
    """
    by = {}
    pa = []
    for r in rows:
        cv = str(r.get("CodigoVolume") or "").strip()
        d = r.get("Data")
        d_str = ""
        if isinstance(d, datetime):
            d_str = d.date().isoformat()
        elif isinstance(d, str):
            d_str = d[:10]
        if de and d_str and d_str < de:
            continue
        if ate and d_str and d_str > ate:
            continue
        v = parse_val_br(r.get("ValorPostagem"))
        if not cv or "_" not in cv:
            # PA VESTI - postagem avulsa sem pedido vinculado
            if v is not None and (r.get("Destinatario") or r.get("CodigoInterno")):
                pa.append({
                    "data": d_str,
                    "operador": r.get("Operador") or "",
                    "modalidade": r.get("Modalidade") or "",
                    "codigoInterno": r.get("CodigoInterno") or "",
                    "numeroNF": str(r.get("NumeroNF") or ""),
                    "remetente": r.get("Remetente") or "",
                    "cliente": r.get("Destinatario") or "",
                    "clienteDoc": str(r.get("CpfCnpjDestinatario") or "").strip(),
                    "cidade": r.get("CidadeDestinatario") or "",
                    "uf": r.get("UFDestinatario") or "",
                    "status": r.get("Status") or "",
                    "postagem": round(v, 2),
                })
            continue
        if cv not in by:
            dom, order = cv.split("_", 1)
            by[cv] = {
                "codigoVolume": cv,
                "orderNumber": r.get("NumeroPedido") or order,
                "domainId": dom,
                "cliente": r.get("Destinatario") or "",
                "clienteDoc": str(r.get("CpfCnpjDestinatario") or "").strip(),
                "cidade": r.get("CidadeDestinatario") or "",
                "uf": r.get("UFDestinatario") or "",
                "status": r.get("Status") or "",
                "data": d_str,
                "postagem": 0.0,
                "remetente": r.get("Remetente") or "",
                "valorDeclarado": parse_val_br(r.get("ValorDeclarado")),
            }
        if v is not None:
            by[cv]["postagem"] += v
    return by, pa


def filter_fabric(pedidos: list[dict], de: str, ate: str) -> dict:
    out = {}
    for p in pedidos:
        d = p.get("data") or ""
        if not d or d < de or d > ate:
            continue
        out[f'{p.get("dominioId","")}_{p.get("orderNumber","")}'] = p
    return out


def fetch_vesti_keys(order_numbers: set[int]) -> set[str]:
    """Existencia autoritativa no Vesti: consulta MongoDB_Pedidos_Geral (base
    COMPLETA, nao so o subconjunto provider=onLog do onlog_data.json) e retorna
    o conjunto de chaves 'domainId_orderNumber' que existem.

    Usado para nao classificar como 'so na planilha' pedido que existe no Vesti
    apenas porque o provider dele nao esta tagueado como onLog no Mongo.

    Falha de Fabric e' NAO-FATAL: retorna set() e o ingest segue (sem a validacao
    extra), so loga aviso.
    """
    nums = sorted({int(n) for n in order_numbers if str(n).strip().isdigit()})
    if not nums:
        return set()
    try:
        from fetch_fabric import connect, load_config
        conn = connect(load_config())
        cur = conn.cursor()
        keys: set[str] = set()
        # consulta em lotes (IN list) p/ nao estourar limite de parametros
        for i in range(0, len(nums), 800):
            chunk = nums[i:i + 800]
            inlist = ",".join(str(n) for n in chunk)
            cur.execute(
                f"SELECT domainId, orderNumber FROM dbo.MongoDB_Pedidos_Geral "
                f"WHERE orderNumber IN ({inlist})"
            )
            for dom, on in cur.fetchall():
                keys.add(f"{dom}_{on}")
        conn.close()
        print(f"      [fabric] base completa: {len(keys)} chaves dominio_pedido existentes no Vesti")
        return keys
    except BaseException as e:
        print(f"      [fabric] AVISO: nao deu p/ validar contra a base completa ({e}). "
              f"Seguindo sem essa validacao (pode haver falso 'so na planilha').")
        return set()


def _is_no_postavel(p: dict) -> bool:
    """Pedidos cancelados ou ainda em SEPARATED nao deveriam aparecer na planilha
    do Diogo - sao 'so no Fabric' esperado e nao representam problema."""
    if p.get("cancelado"):
        return True
    return (p.get("status") or "").upper() == "SEPARATED"


def patch_onlog_data(onlog_data: dict, planilha: dict, de: str, ate: str) -> tuple[int, int]:
    """Atualiza valorPostagem, data (= data de postagem do Diogo), status e
    margemOnlog dos pedidos no onlog_data.json usando os valores reais da planilha
    do Diogo.

    Casamento pela chave dominioId_orderNumber (sem restringir por data Mongo),
    porque o orderDate Mongo as vezes diverge da data real de postagem - ex.:
    cliente faz pedido dia 15 e Diogo posta dia 16. Quem manda na conferencia
    Onlog e' a data da planilha do Diogo (postagem real). Filtramos so' por
    [de, ate] usando a data da planilha pra nao reescrever pedidos de outra
    quinzena ingerida em outro upload.

    Margem = Cotacao BIA - Valor Postagem (lucro real da Vesti por frete).

    Retorna (n_atualizados, n_no_range_sem_planilha).
    """
    # indexa onlog_data por chave domain_order p/ lookup O(1)
    by_key = {f'{p.get("dominioId","")}_{p.get("orderNumber","")}': p
              for p in onlog_data.get("pedidos", [])}
    n_upd = 0
    n_skip = 0
    for k, pl in planilha.items():
        pl_data = pl.get("data") or ""
        if not pl_data or pl_data < de or pl_data > ate:
            continue
        p = by_key.get(k)
        if not p:
            n_skip += 1
            continue
        post = round(pl["postagem"], 2)
        p["valorPostagem"] = post
        p["postagemFonte"] = "planilha-diogo"
        # Sobrescreve a data Mongo (orderDate) pela data real de postagem do Diogo.
        # Isso faz o painel filtrar o pedido pela quinzena CORRETA (a da postagem).
        p["data"] = pl_data
        # Status REAL da postagem (Diogo) - sobrepoe o status do Mongo no display
        if pl.get("status"):
            p["statusOnlog"] = pl["status"]
        bia = p.get("cotacaoBia")
        bia_f = float(bia) if bia is not None else None
        # Margem = Cotacao BIA - Valor Postagem (lucro real)
        if bia_f is not None and bia_f > 0:
            p["margemOnlog"] = round(bia_f - post, 2)
        else:
            # BIA zerado/ausente -> margem nao calculavel (frete gratis ou erro de cadastro)
            p["margemOnlog"] = None
        # Valor Ana FINAL = MAX(Cotacao BIA, Valor Postagem * 1.10)
        if bia_f is not None and bia_f > 0:
            p["valorAnaFinal"] = round(max(bia_f, post * 1.10), 2)
        else:
            p["valorAnaFinal"] = round(post * 1.10, 2)
        n_upd += 1
    return n_upd, n_skip


def fmt_brl(v) -> str:
    if v is None:
        return "-"
    return "R$ " + f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def compare(planilha: dict, fabric: dict, vesti_exists: set[str] | None = None) -> tuple[int, list, list, list]:
    keys = set(planilha) | set(fabric)
    dif, only_p, only_f, ok = [], [], [], 0
    reclass = 0
    for k in keys:
        pl = planilha.get(k)
        fa = fabric.get(k)
        if pl and not fa:
            # Existe no Vesti (base completa do Mongo), so nao estava no onlog_data
            # (provider nao tagueado como onLog) -> NAO e' 'so na planilha'.
            if vesti_exists and k in vesti_exists:
                reclass += 1
                ok += 1
                continue
            only_p.append(pl)
            continue
        if fa and not pl:
            # cancelados e SEPARATED nao deveriam estar na planilha por design - ignora
            if _is_no_postavel(fa):
                continue
            only_f.append(fa)
            continue
        # Ignoramos cliente/destino - sao diferencas de formatacao/abreviacao
        # que nao representam erro real. Comparamos so o que importa: cancelamento e valor postagem.
        divs = []
        if fa.get("cancelado"):
            divs.append(("Cancelado", "(postado pela Onlog)", "CANCELADO no Vesti"))
        pp = pl["postagem"]
        pf = fa.get("valorPostagem")
        if pp is not None and pf is not None:
            if abs(pp - pf) > 0.01:
                divs.append(("Valor Postagem", fmt_brl(pp), fmt_brl(pf)))
        elif pp is not None and pf is None:
            divs.append(("Valor Postagem", fmt_brl(pp), "(sem dado no Fabric)"))
        if divs:
            for campo, a, b in divs:
                dif.append({
                    "orderNumber": fa.get("orderNumber"),
                    "marca": fa.get("marca", "-"),
                    "campo": campo,
                    "planilha": a,
                    "fabric": b,
                })
        else:
            ok += 1
    if reclass:
        print(f"      [fabric] {reclass} reclassificados: existem no Vesti (base completa) -> NAO sao 'so planilha'")
    return ok, dif, only_p, only_f


def process_quinzena(raw: list[dict], de: str, ate: str, onlog_data: dict, xlsx_path: Path, diogo_total: float | None) -> dict:
    """Processa UMA quinzena: aggregate -> merge com snapshot -> patch onlog_data ->
    compare -> escreve onlog_diff_<de>_<ate>.json. Retorna o dict de saida.

    onlog_data e' mutado in-place (patches acumulam). Quem chama escreve no disco.
    """
    print(f"\n=== Quinzena {de} a {ate} ===")
    print(f"[2/4] Agregando planilha (CodigoVolume)")
    planilha, pa_vesti = aggregate_planilha(raw, de, ate)
    print(f"      {len(planilha)} pedidos novos com CodigoVolume; {len(pa_vesti)} PA VESTI novos")

    # Merge com snapshot anterior (se existir) - permite subir planilhas parciais (ex: 1 dia)
    # sem perder dados ja ingeridos da mesma quinzena.
    prev_path = ROOT / f"onlog_diff_{de}_{ate}.json"
    if prev_path.exists():
        try:
            prev = json.loads(prev_path.read_text(encoding="utf-8"))
            snap = prev.get("_planilhaSnapshot") or {}
            old_p = snap.get("planilha") or {}
            old_pa = snap.get("paVesti") or []
            # Bootstrap: se nao tem snapshot mas o onlog_data.json ja tem patches da mesma
            # quinzena (postagemFonte=planilha-diogo), reconstroi a partir deles.
            if not old_p:
                onlog_data_tmp = json.loads(ONLOG_JSON.read_text(encoding="utf-8"))
                for p in onlog_data_tmp.get("pedidos", []):
                    d = p.get("data") or ""
                    if not d or d < de or d > ate:
                        continue
                    if p.get("postagemFonte") != "planilha-diogo":
                        continue
                    cv = f'{p.get("dominioId","")}_{p.get("orderNumber","")}'
                    old_p[cv] = {
                        "codigoVolume": cv,
                        "orderNumber": p.get("orderNumber"),
                        "domainId": p.get("dominioId", ""),
                        "cliente": p.get("cliente", ""),
                        "cidade": p.get("cidade", ""),
                        "uf": p.get("uf", ""),
                        "status": p.get("statusOnlog", ""),
                        "data": d,
                        "postagem": p.get("valorPostagem", 0.0) or 0.0,
                    }
                # Bootstrap PA VESTI a partir do diff antigo
                if not old_pa:
                    old_pa = prev.get("paVesti") or []
                if old_p or old_pa:
                    print(f"      [merge/bootstrap] reconstruido {len(old_p)} pedidos + {len(old_pa)} PA do estado existente")
            # Merge: novos sobrescrevem antigos pela mesma chave
            merged_p = dict(old_p)
            merged_p.update(planilha)
            # PA VESTI: dedup por codigoInterno
            seen = {p.get("codigoInterno") or f"{p.get('data','')}_{p.get('numeroNF','')}_{p.get('postagem',0)}" for p in pa_vesti}
            merged_pa = list(pa_vesti)
            for p in old_pa:
                k = p.get("codigoInterno") or f"{p.get('data','')}_{p.get('numeroNF','')}_{p.get('postagem',0)}"
                if k not in seen:
                    merged_pa.append(p)
                    seen.add(k)
            print(f"      [merge] snapshot anterior: {len(old_p)} pedidos + {len(old_pa)} PA -> total apos merge: {len(merged_p)} + {len(merged_pa)}")
            planilha = merged_p
            pa_vesti = merged_pa
        except Exception as e:
            print(f"      [merge] aviso: nao foi possivel ler snapshot anterior ({e}). Seguindo sem merge.")

    pa_total = round(sum(p["postagem"] for p in pa_vesti), 2)
    print(f"      {len(planilha)} pedidos finais com CodigoVolume")
    print(f"      {len(pa_vesti)} postagens avulsas PA VESTI (sem pedido) - total R$ {pa_total:,.2f}")

    print(f"[3/4] Filtrando Fabric")
    # Filtra Fabric pelo range REAL coberto pela planilha (e nao pela quinzena
    # inteira), pra evitar falso "so Fabric" quando o Diogo ainda nao exportou
    # todos os dias. Clampa dentro de [de, ate] da quinzena.
    datas_pl = [p.get("data") for p in planilha.values() if p.get("data")]
    if datas_pl:
        pl_de = max(de, min(datas_pl))
        pl_ate = min(ate, max(datas_pl))
    else:
        pl_de, pl_ate = de, ate
    fabric = filter_fabric(onlog_data.get("pedidos", []), pl_de, pl_ate)
    print(f"      {len(fabric)} pedidos do Fabric no range coberto pela planilha ({pl_de}..{pl_ate})")

    print(f"[3.5/4] Patch onlog_data com valores da planilha (postagem + margem)")
    n_upd, n_skip = patch_onlog_data(onlog_data, planilha, de, ate)
    print(f"      {n_upd} pedidos atualizados (postagem + margem); {n_skip} sem match na planilha")

    print(f"[3.8/4] Validando contra onlog_data (pedidos com provider=onLog)")
    # Antes consultavamos MongoDB_Pedidos_Geral (base ampla), mas isso fazia sumir
    # pedidos cujo provider no Mongo nao esta como onLog: caiam fora de 'so planilha'
    # via reclass MAS nao apareciam em onlog_data => limbo. Agora reclassificamos
    # somente quando o pedido existe no proprio onlog_data (que e' o que o painel
    # exibe). Casos de dominio divergente dentro de onLog continuam sendo capturados
    # pelo _vestiMatch do template (orderNumber + CPF).
    vesti_exists = {
        f'{p.get("dominioId","")}_{p.get("orderNumber","")}'
        for p in onlog_data.get("pedidos", [])
        if p.get("orderNumber") is not None
    }
    print(f"      onlog_data: {len(vesti_exists)} chaves dominio_pedido (provider=onLog)")

    print(f"[4/4] Comparando")
    ok, dif, only_p, only_f = compare(planilha, fabric, vesti_exists)

    # Auditoria simplificada: 3 tipos de problema relevantes
    audit_status, audit_etiqueta, audit_frete = [], [], []
    THRESHOLD_FRETE_PCT = 0.05  # 5% de diferença
    for k, pl in planilha.items():
        fa = fabric.get(k)
        if not fa or _is_no_postavel(fa):
            continue
        # 1) Status divergente (Mongo nao acompanha o status real da planilha)
        st_mongo = (fa.get("status") or "").upper()
        st_pla = (pl.get("status") or "").upper()
        # Mongo: WAITING/SEPARATED ainda nao postados; SENT em transito; DELIVERED entregue
        # Planilha: ENTREGUE / EM TRANSITO / etc - se planilha esta avancada e Mongo atras, divergencia
        mongo_atras = st_mongo in ("WAITING", "SEPARATED") and st_pla and "RETIRADA" not in st_pla
        if mongo_atras:
            audit_status.append({
                "orderNumber": fa.get("orderNumber"),
                "marca": fa.get("marca", "-"),
                "cliente": fa.get("cliente", "-"),
                "statusMongo": fa.get("status") or "-",
                "statusOnlog": pl.get("status") or "-",
            })
        # 2) Etiqueta ausente no Mongo (mas pedido foi postado pelo Diogo)
        if not fa.get("comEtiqueta"):
            audit_etiqueta.append({
                "orderNumber": fa.get("orderNumber"),
                "marca": fa.get("marca", "-"),
                "cliente": fa.get("cliente", "-"),
                "data": fa.get("data"),
                "statusMongo": fa.get("status") or "-",
                "statusOnlog": pl.get("status") or "-",
                "postagem": round(pl["postagem"], 2),
            })
        # 3) Diferenca de frete relevante (BIA Mongo vs Postagem*1.10)
        post = pl["postagem"]
        bia = fa.get("cotacaoBia")
        if post and bia and bia > 0:
            esperado = post * 1.10
            diff_pct = abs(bia - esperado) / esperado if esperado else 0
            if diff_pct > THRESHOLD_FRETE_PCT and abs(bia - esperado) > 1.0:
                audit_frete.append({
                    "orderNumber": fa.get("orderNumber"),
                    "marca": fa.get("marca", "-"),
                    "cliente": fa.get("cliente", "-"),
                    "postagem": round(post, 2),
                    "esperado": round(esperado, 2),
                    "biaMongo": round(bia, 2),
                    "diferenca": round(bia - esperado, 2),
                    "diferencaPct": round(diff_pct * 100, 1),
                })

    print(f"      [auditoria] status atrasado Mongo: {len(audit_status)}")
    print(f"      [auditoria] sem etiqueta Mongo:    {len(audit_etiqueta)}")
    print(f"      [auditoria] frete divergente >5%:  {len(audit_frete)}")
    n_dif_uniq = len({d["orderNumber"] for d in dif})
    print(f"      OK={ok}  Divergencias={len(dif)} ({n_dif_uniq} pedidos)")
    print(f"      So planilha={len(only_p)}  So Fabric={len(only_f)}")

    # totais financeiros
    total_pedidos_planilha = round(sum(p["postagem"] for p in planilha.values()), 2)
    # cobranca por marca (postagem * 1.10) - so para pedidos com domainId valido
    cobranca_marca = {}
    for p in planilha.values():
        dom = p["domainId"]
        cobranca_marca.setdefault(dom, {"domainId": dom, "nPedidos": 0, "postagem": 0.0, "cobrar": 0.0})
        cobranca_marca[dom]["nPedidos"] += 1
        cobranca_marca[dom]["postagem"] += p["postagem"]
        cobranca_marca[dom]["cobrar"] += p["postagem"] * 1.10
    # Anexar nome da marca via companies_data.json
    companies_path = ROOT / "companies_data.json"
    nome_por_dom = {}
    if companies_path.exists():
        cs = json.loads(companies_path.read_text(encoding="utf-8"))
        for c in cs:
            d = str(c.get("domain_id") or "")
            if not d:
                continue
            if c.get("isMatriz") or d not in nome_por_dom:
                nome_por_dom[d] = c.get("nome_fantasia") or c.get("name") or ""
        for dom, info in cobranca_marca.items():
            info["marca"] = nome_por_dom.get(dom, "")
    cobranca_lista = sorted(
        [{"domainId": d, "marca": v.get("marca", ""), "nPedidos": v["nPedidos"],
          "postagem": round(v["postagem"], 2), "cobrar": round(v["cobrar"], 2)}
         for d, v in cobranca_marca.items()],
        key=lambda x: -x["cobrar"]
    )

    total_geral = round(total_pedidos_planilha + pa_total, 2)
    # Custo real = o que o Diogo cobra (linha totalizadora). Fallback: soma da planilha.
    custo_diogo = round(diogo_total, 2) if diogo_total is not None else total_geral
    # Receita = custo Diogo * 1.10 (Vesti cobra das marcas o que paga + 10%)
    total_cobrar = round(custo_diogo * 1.10, 2)
    out = {
        "quinzena": {"de": de, "ate": ate},
        "rangeCoberto": {"de": pl_de, "ate": pl_ate},
        "geradoEm": datetime.now().isoformat(),
        "planilhaArquivo": xlsx_path.name,
        "resumo": {
            "ok": ok,
            "divergencias": len(dif),
            "divergenciasPedidos": n_dif_uniq,
            "soPlanilha": len(only_p),
            "soFabric": len(only_f),
            "nPlanilha": len(planilha),
            "nFabric": len(fabric),
            "totalPedidosPostagem": total_pedidos_planilha,
            "totalPaVesti": pa_total,
            "totalGeralPostagem": total_geral,
            "totalCobrarPedidos": round(total_pedidos_planilha * 1.10, 2),
            "totalCobrarPa": round(pa_total * 1.10, 2),
            "totalCobrarMarcas": total_cobrar,
            "custoDiogo": custo_diogo,
            "diogoDetectado": diogo_total is not None,
            "margemVesti": round(total_cobrar - custo_diogo, 2),
            "nPaVesti": len(pa_vesti),
        },
        "paVesti": pa_vesti,
        "cobrancaPorMarca": cobranca_lista,
        "auditStatus": audit_status,
        "auditEtiqueta": audit_etiqueta,
        "auditFrete": audit_frete,
        "auditResumo": {
            "statusAtrasado": len(audit_status),
            "semEtiqueta": len(audit_etiqueta),
            "freteDivergente": len(audit_frete),
        },
        "divergencias": dif,
        "soPlanilha": [{
            "codigoVolume": p["codigoVolume"],
            "orderNumber": p["orderNumber"],
            "cliente": p["cliente"],
            "clienteDoc": p.get("clienteDoc", ""),
            "cidade": p["cidade"],
            "uf": p["uf"],
            "status": p["status"],
            "postagem": round(p["postagem"], 2),
            "data": p.get("data", ""),
            # Usa marca de companies_data (mesma string que o painel mostra p/ outros
            # pedidos da mesma marca) e cai pro remetente da planilha so se nao tiver
            # match. Garante que pesquisa por marca encontra os pedidos so-planilha.
            "remetente": nome_por_dom.get(p.get("domainId", "")) or p.get("remetente", ""),
            "valorDeclarado": p.get("valorDeclarado"),
        } for p in only_p],
        "soFabric": [{
            "data": p.get("data"),
            "orderNumber": p.get("orderNumber"),
            "marca": p.get("marca", "-"),
            "cliente": p.get("cliente", "-"),
            "cidade": p.get("cidade", ""),
            "uf": p.get("uf", ""),
            "cancelado": bool(p.get("cancelado")),
            "status": p.get("status", ""),
            "valor": p.get("valor"),
        } for p in only_f],
        # Snapshot para permitir merge incremental em re-ingestoes da mesma quinzena
        "_planilhaSnapshot": {
            "planilha": planilha,
            "paVesti": pa_vesti,
        },
    }
    # Salva por quinzena: onlog_diff_<de>_<ate>.json (multi-quinzenas)
    out_quinzena = ROOT / f"onlog_diff_{de}_{ate}.json"
    out_quinzena.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f">> {out_quinzena.name} escrito ({out_quinzena.stat().st_size//1024} KB)")
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="Caminho da planilha do Diogo")
    ap.add_argument("--de", help="Data inicial (default: detectado)")
    ap.add_argument("--ate", help="Data final (default: detectado)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"ERRO: nao achei {xlsx_path}")
        sys.exit(1)
    if not ONLOG_JSON.exists():
        print(f"ERRO: {ONLOG_JSON} nao existe. Rode py fetch_onlog.py antes.")
        sys.exit(1)

    print(f"[1/4] Lendo planilha: {xlsx_path.name}")
    raw = read_planilha(xlsx_path)
    print(f"      {len(raw)} linhas brutas")
    diogo_total = detect_diogo_total(raw)
    if diogo_total is not None:
        print(f"      [linha totalizadora detectada] Diogo cobra: R$ {diogo_total:,.2f}")

    if args.de and args.ate:
        ranges = [(args.de, args.ate)]
    else:
        ranges = detect_quinzenas(raw)
        if not ranges:
            print("ERRO: nao consegui detectar nenhuma quinzena. Use --de/--ate.")
            sys.exit(1)
    print(f"      quinzenas detectadas: {', '.join(f'{a}..{b}' for a,b in ranges)}")

    onlog_data = json.loads(ONLOG_JSON.read_text(encoding="utf-8"))
    last_out = None
    for de, ate in ranges:
        last_out = process_quinzena(raw, de, ate, onlog_data, xlsx_path, diogo_total)
    # Escreve onlog_data uma vez no final (com patches acumulados de todas as quinzenas)
    ONLOG_JSON.write_text(json.dumps(onlog_data, ensure_ascii=False), encoding="utf-8")
    # Mantem onlog_diff.json como compat (ultima quinzena processada)
    if last_out is not None:
        (ROOT / "onlog_diff.json").write_text(json.dumps(last_out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n>> {len(ranges)} quinzena(s) processada(s). Agora rode: py merge_data.py && py build_html.py")


if __name__ == "__main__":
    main()
